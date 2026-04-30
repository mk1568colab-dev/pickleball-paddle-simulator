"""Run 20 named offline market environments and export one Excel workbook.

This script is intentionally separate from the Streamlit app. It uses the real
simulator engine in memory and does not read or write the classroom SQLite
database.
"""

from __future__ import annotations

import argparse
import csv
import math
import sys
from collections import Counter, defaultdict
from copy import copy, deepcopy
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError as error:  # pragma: no cover - user environment guard
    raise SystemExit(
        "Missing Excel dependency. Run `python -m pip install -r requirements.txt` "
        "and try again. This script requires `openpyxl`."
    ) from error


ROOT_DIR = Path(__file__).resolve().parents[1]
SCRIPT_DIR = Path(__file__).resolve().parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import run_strategy_simulation as runner


STRATEGY_ROTATION = (
    "cash_conservative",
    "balanced_sop",
    "premium_quality",
    "innovation_leap",
    "aggressive_growth",
    "low_cost_volume",
)

BASE_MARKET_FOR_ROUND = runner.market_for_round
BASE_DEMAND_GROWTH_RATE = 0.07
OUTPUT_ROOT = ROOT_DIR / "simulation_outputs"
RANDOM_SEED = 20260428
MODE_BASELINE = "baseline"
MODE_CALIBRATED = "calibrated"


SCENARIO_DEFINITIONS: list[dict[str, Any]] = [
    {
        "scenario_id": 1,
        "sheet_name": "01_Baseline",
        "key": "baseline_20",
        "label": "Baseline",
        "represents": "Balanced demand, normal cost, and moderate technology pressure.",
        "learning": "Use this as the neutral reference case before introducing shocks.",
        "discussion_question": "Which strategy performs well without needing a special market advantage?",
        "demand_multiplier": 1.00,
        "demand_growth_delta": 0.00,
        "premium_share_shift": 0.00,
        "mid_share_shift": 0.00,
        "beginner_share_shift": 0.00,
        "material_cost_index_delta": 0.00,
        "supply_risk_override": None,
        "quality_sensitivity_delta": 0.00,
        "technology_shift_delta": 0.00,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.00,
        "mid_tech_adoption_delta": 0.00,
        "beginner_price_pressure_delta": 0.00,
        "price_sensitivity": 1.00,
        "forecast_volatility": 0.00,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.00,
    },
    {
        "scenario_id": 2,
        "sheet_name": "02_Picky_Customers",
        "key": "picky_customers",
        "label": "Picky Customers",
        "represents": "Customers are selective and punish poor quality or weak service.",
        "learning": "Quality, service, and reputation can matter as much as price.",
        "discussion_question": "When does spending on QC become an investment rather than a cost?",
        "demand_multiplier": 0.98,
        "demand_growth_delta": 0.01,
        "premium_share_shift": 0.04,
        "mid_share_shift": 0.03,
        "beginner_share_shift": -0.07,
        "material_cost_index_delta": 0.03,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.22,
        "technology_shift_delta": 0.03,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.05,
        "mid_tech_adoption_delta": 0.04,
        "beginner_price_pressure_delta": -0.08,
        "price_sensitivity": 0.85,
        "forecast_volatility": 0.03,
        "customer_pickiness": 1.30,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.04,
    },
    {
        "scenario_id": 3,
        "sheet_name": "03_Price_Sensitive",
        "key": "price_sensitive",
        "label": "Price Sensitive",
        "represents": "Buyers trade down and compare prices aggressively.",
        "learning": "Low prices only help if margins, defects, and inventory are controlled.",
        "discussion_question": "Which costs must a low-price team control before volume becomes profitable?",
        "demand_multiplier": 1.10,
        "demand_growth_delta": 0.02,
        "premium_share_shift": -0.08,
        "mid_share_shift": -0.02,
        "beginner_share_shift": 0.10,
        "material_cost_index_delta": -0.02,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": -0.12,
        "technology_shift_delta": -0.02,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": -0.08,
        "mid_tech_adoption_delta": -0.05,
        "beginner_price_pressure_delta": 0.28,
        "price_sensitivity": 1.45,
        "forecast_volatility": 0.04,
        "customer_pickiness": 0.82,
        "price_war_pressure": 0.45,
        "cash_stress_factor": 0.98,
    },
    {
        "scenario_id": 4,
        "sheet_name": "04_Tech_Shift",
        "key": "tech_shift_20",
        "label": "Tech Shift",
        "represents": "Customers move toward newer paddle technology generations.",
        "learning": "Innovation timing and product retirement become more important.",
        "discussion_question": "How early should a team invest before current products become obsolete?",
        "demand_multiplier": 1.08,
        "demand_growth_delta": 0.03,
        "premium_share_shift": 0.08,
        "mid_share_shift": 0.03,
        "beginner_share_shift": -0.11,
        "material_cost_index_delta": 0.04,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.06,
        "technology_shift_delta": 0.34,
        "market_generation_offset": 1,
        "premium_tech_adoption_delta": 0.24,
        "mid_tech_adoption_delta": 0.18,
        "beginner_price_pressure_delta": 0.00,
        "price_sensitivity": 0.95,
        "forecast_volatility": 0.05,
        "customer_pickiness": 1.05,
        "price_war_pressure": 0.05,
        "cash_stress_factor": 1.06,
    },
    {
        "scenario_id": 5,
        "sheet_name": "05_Beginner_Boom",
        "key": "beginner_boom",
        "label": "Beginner Boom",
        "represents": "New players enter the market and beginner demand expands quickly.",
        "learning": "Volume opportunities still require margin and capacity discipline.",
        "discussion_question": "Why might a high-volume team still lose money during a demand boom?",
        "demand_multiplier": 1.40,
        "demand_growth_delta": 0.08,
        "premium_share_shift": -0.10,
        "mid_share_shift": -0.03,
        "beginner_share_shift": 0.13,
        "material_cost_index_delta": 0.00,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": -0.10,
        "technology_shift_delta": -0.02,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": -0.10,
        "mid_tech_adoption_delta": -0.04,
        "beginner_price_pressure_delta": 0.22,
        "price_sensitivity": 1.25,
        "forecast_volatility": 0.08,
        "customer_pickiness": 0.88,
        "price_war_pressure": 0.25,
        "cash_stress_factor": 1.00,
    },
    {
        "scenario_id": 6,
        "sheet_name": "06_Premium_Market",
        "key": "premium_market",
        "label": "Premium Market",
        "represents": "Premium buyers expand and accept higher prices for quality and technology.",
        "learning": "Premium positioning can work when demand supports margin and reputation.",
        "discussion_question": "What must a premium team do operationally to defend high prices?",
        "demand_multiplier": 1.12,
        "demand_growth_delta": 0.04,
        "premium_share_shift": 0.16,
        "mid_share_shift": -0.03,
        "beginner_share_shift": -0.13,
        "material_cost_index_delta": 0.06,
        "supply_risk_override": "Low",
        "quality_sensitivity_delta": 0.18,
        "technology_shift_delta": 0.12,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.18,
        "mid_tech_adoption_delta": 0.08,
        "beginner_price_pressure_delta": -0.12,
        "price_sensitivity": 0.72,
        "forecast_volatility": 0.04,
        "customer_pickiness": 1.22,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.05,
    },
    {
        "scenario_id": 7,
        "sheet_name": "07_Demand_Recession",
        "key": "demand_recession",
        "label": "Demand Recession",
        "represents": "Market demand contracts and growth turns negative.",
        "learning": "Forecast discipline and inventory control matter when demand shrinks.",
        "discussion_question": "Which teams were stuck with costs from plans made for a bigger market?",
        "demand_multiplier": 0.72,
        "demand_growth_delta": -0.08,
        "premium_share_shift": -0.06,
        "mid_share_shift": -0.02,
        "beginner_share_shift": 0.08,
        "material_cost_index_delta": 0.02,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": -0.04,
        "technology_shift_delta": -0.02,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": -0.04,
        "mid_tech_adoption_delta": -0.03,
        "beginner_price_pressure_delta": 0.16,
        "price_sensitivity": 1.20,
        "forecast_volatility": 0.07,
        "customer_pickiness": 0.92,
        "price_war_pressure": 0.20,
        "cash_stress_factor": 1.08,
    },
    {
        "scenario_id": 8,
        "sheet_name": "08_Demand_Boom",
        "key": "demand_boom_20",
        "label": "Demand Boom",
        "represents": "Demand grows quickly across the market.",
        "learning": "Growth creates opportunity but also capacity, inventory, and cash pressure.",
        "discussion_question": "Which constraint prevented teams from converting demand into profit?",
        "demand_multiplier": 1.55,
        "demand_growth_delta": 0.10,
        "premium_share_shift": -0.01,
        "mid_share_shift": 0.03,
        "beginner_share_shift": -0.02,
        "material_cost_index_delta": 0.03,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.02,
        "technology_shift_delta": 0.04,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.04,
        "mid_tech_adoption_delta": 0.03,
        "beginner_price_pressure_delta": -0.04,
        "price_sensitivity": 0.90,
        "forecast_volatility": 0.06,
        "customer_pickiness": 0.98,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.04,
    },
    {
        "scenario_id": 9,
        "sheet_name": "09_Supply_Shock",
        "key": "supply_shock_20",
        "label": "Supply Shock",
        "represents": "Supply becomes unreliable and materials become more expensive.",
        "learning": "Supplier mix and raw-material planning affect both service and defects.",
        "discussion_question": "Did teams protect service by paying more, or protect cost by taking supply risk?",
        "demand_multiplier": 0.98,
        "demand_growth_delta": -0.01,
        "premium_share_shift": 0.00,
        "mid_share_shift": -0.02,
        "beginner_share_shift": 0.02,
        "material_cost_index_delta": 0.18,
        "supply_risk_override": "High",
        "quality_sensitivity_delta": 0.08,
        "technology_shift_delta": 0.01,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.00,
        "mid_tech_adoption_delta": 0.00,
        "beginner_price_pressure_delta": 0.08,
        "price_sensitivity": 1.05,
        "forecast_volatility": 0.06,
        "customer_pickiness": 1.05,
        "price_war_pressure": 0.05,
        "cash_stress_factor": 1.18,
    },
    {
        "scenario_id": 10,
        "sheet_name": "10_Cost_Inflation",
        "key": "cost_inflation",
        "label": "Cost Inflation",
        "represents": "Input costs rise even without severe supply disruption.",
        "learning": "Pricing, sourcing, and margin discipline become central.",
        "discussion_question": "Which teams protected gross margin without destroying demand?",
        "demand_multiplier": 1.00,
        "demand_growth_delta": 0.01,
        "premium_share_shift": 0.01,
        "mid_share_shift": 0.00,
        "beginner_share_shift": -0.01,
        "material_cost_index_delta": 0.24,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.02,
        "technology_shift_delta": 0.02,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.02,
        "mid_tech_adoption_delta": 0.01,
        "beginner_price_pressure_delta": 0.10,
        "price_sensitivity": 1.12,
        "forecast_volatility": 0.03,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.08,
        "cash_stress_factor": 1.22,
    },
    {
        "scenario_id": 11,
        "sheet_name": "11_Fast_Lifecycle",
        "key": "fast_lifecycle",
        "label": "Fast Lifecycle",
        "represents": "Products age quickly because customer expectations shift.",
        "learning": "Portfolio renewal and retirement timing become more important.",
        "discussion_question": "When should a team stop supporting a mature product and replace it?",
        "demand_multiplier": 1.05,
        "demand_growth_delta": 0.02,
        "premium_share_shift": 0.05,
        "mid_share_shift": 0.02,
        "beginner_share_shift": -0.07,
        "material_cost_index_delta": 0.05,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.10,
        "technology_shift_delta": 0.30,
        "market_generation_offset": 1,
        "premium_tech_adoption_delta": 0.18,
        "mid_tech_adoption_delta": 0.16,
        "beginner_price_pressure_delta": 0.02,
        "price_sensitivity": 0.96,
        "forecast_volatility": 0.08,
        "customer_pickiness": 1.10,
        "price_war_pressure": 0.08,
        "cash_stress_factor": 1.08,
    },
    {
        "scenario_id": 12,
        "sheet_name": "12_Warranty_Sensitive",
        "key": "warranty_sensitive",
        "label": "Warranty Sensitive",
        "represents": "Defects become especially costly through reputation and warranty exposure.",
        "learning": "Quality failures create financial and market consequences.",
        "discussion_question": "Which teams reduced defects enough to justify higher QC or supplier costs?",
        "demand_multiplier": 1.00,
        "demand_growth_delta": 0.00,
        "premium_share_shift": 0.06,
        "mid_share_shift": 0.02,
        "beginner_share_shift": -0.08,
        "material_cost_index_delta": 0.07,
        "supply_risk_override": "High",
        "quality_sensitivity_delta": 0.30,
        "technology_shift_delta": 0.05,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.08,
        "mid_tech_adoption_delta": 0.06,
        "beginner_price_pressure_delta": -0.06,
        "price_sensitivity": 0.86,
        "forecast_volatility": 0.04,
        "customer_pickiness": 1.35,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.10,
    },
    {
        "scenario_id": 13,
        "sheet_name": "13_Price_War",
        "key": "price_war",
        "label": "Price War",
        "represents": "Competitors chase price-sensitive beginner and mid-market buyers.",
        "learning": "Winning volume is not the same as winning profit.",
        "discussion_question": "How much price pressure can a team absorb before the business breaks?",
        "demand_multiplier": 1.25,
        "demand_growth_delta": 0.04,
        "premium_share_shift": -0.13,
        "mid_share_shift": -0.01,
        "beginner_share_shift": 0.14,
        "material_cost_index_delta": -0.01,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": -0.16,
        "technology_shift_delta": -0.04,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": -0.10,
        "mid_tech_adoption_delta": -0.06,
        "beginner_price_pressure_delta": 0.38,
        "price_sensitivity": 1.60,
        "forecast_volatility": 0.07,
        "customer_pickiness": 0.78,
        "price_war_pressure": 0.80,
        "cash_stress_factor": 1.00,
    },
    {
        "scenario_id": 14,
        "sheet_name": "14_Inventory_Risk",
        "key": "inventory_risk",
        "label": "Inventory Risk",
        "represents": "Demand is weak and mistakes leave teams with costly inventory.",
        "learning": "Inventory can protect service, but excess inventory consumes cash.",
        "discussion_question": "Which inventories were strategic buffers and which were planning errors?",
        "demand_multiplier": 0.82,
        "demand_growth_delta": -0.04,
        "premium_share_shift": -0.03,
        "mid_share_shift": 0.00,
        "beginner_share_shift": 0.03,
        "material_cost_index_delta": 0.10,
        "supply_risk_override": "Low",
        "quality_sensitivity_delta": 0.00,
        "technology_shift_delta": 0.01,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.00,
        "mid_tech_adoption_delta": 0.00,
        "beginner_price_pressure_delta": 0.12,
        "price_sensitivity": 1.08,
        "forecast_volatility": 0.10,
        "customer_pickiness": 0.95,
        "price_war_pressure": 0.15,
        "cash_stress_factor": 1.12,
    },
    {
        "scenario_id": 15,
        "sheet_name": "15_Capacity_Constraint",
        "key": "capacity_constraint",
        "label": "Capacity Constraint",
        "represents": "Demand is strong but operating constraints limit fulfillment.",
        "learning": "Capacity, overtime, and expansion choices must match demand and cash.",
        "discussion_question": "Which teams lost demand because they did not prepare capacity or materials?",
        "demand_multiplier": 1.50,
        "demand_growth_delta": 0.09,
        "premium_share_shift": 0.00,
        "mid_share_shift": 0.04,
        "beginner_share_shift": -0.04,
        "material_cost_index_delta": 0.08,
        "supply_risk_override": "High",
        "quality_sensitivity_delta": 0.04,
        "technology_shift_delta": 0.03,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.04,
        "mid_tech_adoption_delta": 0.04,
        "beginner_price_pressure_delta": -0.02,
        "price_sensitivity": 0.92,
        "forecast_volatility": 0.05,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.10,
    },
    {
        "scenario_id": 16,
        "sheet_name": "16_Forecast_Volatility",
        "key": "forecast_volatility",
        "label": "Forecast Volatility",
        "represents": "Demand swings around a trend, creating plan-vs-actual errors.",
        "learning": "S&OP discipline is tested when demand is not smooth.",
        "discussion_question": "How should teams plan when the forecast is directionally right but round demand swings?",
        "demand_multiplier": 1.08,
        "demand_growth_delta": 0.02,
        "premium_share_shift": 0.00,
        "mid_share_shift": 0.03,
        "beginner_share_shift": -0.03,
        "material_cost_index_delta": 0.03,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.04,
        "technology_shift_delta": 0.04,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.03,
        "mid_tech_adoption_delta": 0.03,
        "beginner_price_pressure_delta": 0.05,
        "price_sensitivity": 1.00,
        "forecast_volatility": 0.22,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.05,
        "cash_stress_factor": 1.06,
    },
    {
        "scenario_id": 17,
        "sheet_name": "17_Viral_Demand",
        "key": "viral_demand",
        "label": "Viral Demand",
        "represents": "A sudden late surge creates demand that was difficult to plan for.",
        "learning": "Upside demand can still create lost sales if capacity and materials lag.",
        "discussion_question": "Should teams build slack for upside demand, or protect cash and accept stockouts?",
        "demand_multiplier": 1.18,
        "demand_growth_delta": 0.11,
        "premium_share_shift": 0.02,
        "mid_share_shift": 0.04,
        "beginner_share_shift": -0.06,
        "material_cost_index_delta": 0.04,
        "supply_risk_override": "Moderate",
        "quality_sensitivity_delta": 0.02,
        "technology_shift_delta": 0.08,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.06,
        "mid_tech_adoption_delta": 0.05,
        "beginner_price_pressure_delta": -0.03,
        "price_sensitivity": 0.90,
        "forecast_volatility": 0.18,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.04,
        "late_surge_round": 8,
        "late_surge_multiplier": 1.32,
    },
    {
        "scenario_id": 18,
        "sheet_name": "18_Cash_Crunch",
        "key": "cash_crunch",
        "label": "Cash Crunch",
        "represents": "Cash is stressed by weak demand, higher cost, and supply risk.",
        "learning": "Liquidity pressure can dominate attractive operating plans.",
        "discussion_question": "Which decisions created cash pressure before teams saw the market outcome?",
        "demand_multiplier": 0.88,
        "demand_growth_delta": -0.03,
        "premium_share_shift": -0.02,
        "mid_share_shift": 0.02,
        "beginner_share_shift": 0.00,
        "material_cost_index_delta": 0.16,
        "supply_risk_override": "High",
        "quality_sensitivity_delta": 0.04,
        "technology_shift_delta": 0.02,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.00,
        "mid_tech_adoption_delta": 0.00,
        "beginner_price_pressure_delta": 0.18,
        "price_sensitivity": 1.18,
        "forecast_volatility": 0.09,
        "customer_pickiness": 1.00,
        "price_war_pressure": 0.20,
        "cash_stress_factor": 1.28,
    },
    {
        "scenario_id": 19,
        "sheet_name": "19_Premium_Expansion",
        "key": "premium_expansion",
        "label": "Premium Expansion",
        "represents": "Premium and mid-tier demand expand together.",
        "learning": "Teams can grow profitably if price, quality, and capacity scale together.",
        "discussion_question": "Which team scaled premium demand without sacrificing service or liquidity?",
        "demand_multiplier": 1.32,
        "demand_growth_delta": 0.08,
        "premium_share_shift": 0.12,
        "mid_share_shift": 0.04,
        "beginner_share_shift": -0.16,
        "material_cost_index_delta": 0.05,
        "supply_risk_override": "Low",
        "quality_sensitivity_delta": 0.16,
        "technology_shift_delta": 0.10,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.16,
        "mid_tech_adoption_delta": 0.10,
        "beginner_price_pressure_delta": -0.10,
        "price_sensitivity": 0.78,
        "forecast_volatility": 0.05,
        "customer_pickiness": 1.18,
        "price_war_pressure": 0.00,
        "cash_stress_factor": 1.04,
    },
    {
        "scenario_id": 20,
        "sheet_name": "20_Unstable_Market",
        "key": "unstable_market",
        "label": "Unstable Market",
        "represents": "Demand, costs, quality expectations, and supply risk are all volatile.",
        "learning": "Robust strategies survive better than fragile optimized plans.",
        "discussion_question": "Which strategy was most robust when several assumptions moved at once?",
        "demand_multiplier": 1.02,
        "demand_growth_delta": 0.01,
        "premium_share_shift": 0.03,
        "mid_share_shift": -0.01,
        "beginner_share_shift": -0.02,
        "material_cost_index_delta": 0.10,
        "supply_risk_override": "High",
        "quality_sensitivity_delta": 0.12,
        "technology_shift_delta": 0.16,
        "market_generation_offset": 0,
        "premium_tech_adoption_delta": 0.10,
        "mid_tech_adoption_delta": 0.08,
        "beginner_price_pressure_delta": 0.12,
        "price_sensitivity": 1.10,
        "forecast_volatility": 0.24,
        "customer_pickiness": 1.15,
        "price_war_pressure": 0.20,
        "cash_stress_factor": 1.18,
    },
]


STRATEGY_STRENGTHS = {
    "Cash conservative": "Protects liquidity, avoids debt, and limits self-inflicted operating risk.",
    "Balanced S&OP": "Keeps forecast, production, sourcing, and service decisions relatively aligned.",
    "Premium quality": "Produces strong service/forecast outcomes when customers reward reliability.",
    "Innovation leap": "Builds future technology position and can benefit in fast technology shifts.",
    "Aggressive growth": "Can chase demand upside when capacity and working capital are available.",
    "Low-cost volume": "Can serve high-volume beginner demand when margins remain protected.",
}

STRATEGY_WEAKNESSES = {
    "Cash conservative": "May under-serve demand in boom markets and can look too cautious.",
    "Balanced S&OP": "Can carry enough borrowing or investment burden to weaken liquidity.",
    "Premium quality": "High quality and sourcing costs can overwhelm revenue if demand is not premium enough.",
    "Innovation leap": "Large NPD investment can create debt and weak short-term service.",
    "Aggressive growth": "Expansion, overtime, and inventory can destroy cash if demand or margins disappoint.",
    "Low-cost volume": "Thin margins, defects, and inventory cost can erase volume benefits.",
}


CALIBRATED_SCENARIO_OVERRIDES: dict[str, dict[str, Any]] = {
    "picky_customers": {
        "premium_share_shift": 0.12,
        "mid_share_shift": 0.04,
        "beginner_share_shift": -0.16,
        "quality_sensitivity_delta": 0.36,
        "customer_pickiness": 1.50,
        "material_cost_index_delta": -0.02,
        "supply_risk_override": "Low",
    },
    "price_sensitive": {
        "demand_multiplier": 1.35,
        "premium_share_shift": -0.16,
        "mid_share_shift": -0.02,
        "beginner_share_shift": 0.18,
        "quality_sensitivity_delta": -0.24,
        "beginner_price_pressure_delta": 0.50,
        "price_sensitivity": 1.75,
        "material_cost_index_delta": -0.16,
    },
    "tech_shift_20": {
        "demand_multiplier": 1.22,
        "premium_share_shift": 0.16,
        "mid_share_shift": 0.06,
        "beginner_share_shift": -0.22,
        "technology_shift_delta": 0.72,
        "market_generation_offset": 2,
        "premium_tech_adoption_delta": 0.40,
        "mid_tech_adoption_delta": 0.30,
        "material_cost_index_delta": -0.03,
    },
    "beginner_boom": {
        "demand_multiplier": 1.85,
        "demand_growth_delta": 0.12,
        "premium_share_shift": -0.18,
        "mid_share_shift": -0.04,
        "beginner_share_shift": 0.22,
        "quality_sensitivity_delta": -0.22,
        "beginner_price_pressure_delta": 0.44,
        "material_cost_index_delta": -0.18,
    },
    "premium_market": {
        "demand_multiplier": 1.35,
        "premium_share_shift": 0.26,
        "mid_share_shift": -0.02,
        "beginner_share_shift": -0.24,
        "quality_sensitivity_delta": 0.34,
        "premium_tech_adoption_delta": 0.34,
        "price_sensitivity": 0.58,
        "material_cost_index_delta": -0.06,
    },
    "demand_boom_20": {
        "demand_multiplier": 2.10,
        "demand_growth_delta": 0.16,
        "material_cost_index_delta": -0.06,
        "supply_risk_override": "Low",
        "forecast_volatility": 0.04,
    },
    "fast_lifecycle": {
        "demand_multiplier": 1.22,
        "technology_shift_delta": 0.78,
        "market_generation_offset": 2,
        "premium_tech_adoption_delta": 0.38,
        "mid_tech_adoption_delta": 0.32,
        "material_cost_index_delta": -0.02,
    },
    "warranty_sensitive": {
        "premium_share_shift": 0.16,
        "mid_share_shift": 0.04,
        "beginner_share_shift": -0.20,
        "quality_sensitivity_delta": 0.42,
        "customer_pickiness": 1.55,
        "supply_risk_override": "High",
    },
    "price_war": {
        "demand_multiplier": 1.75,
        "premium_share_shift": -0.22,
        "mid_share_shift": -0.04,
        "beginner_share_shift": 0.26,
        "quality_sensitivity_delta": -0.28,
        "beginner_price_pressure_delta": 0.58,
        "price_sensitivity": 1.90,
        "material_cost_index_delta": -0.20,
    },
    "capacity_constraint": {
        "demand_multiplier": 2.20,
        "demand_growth_delta": 0.18,
        "material_cost_index_delta": -0.08,
        "supply_risk_override": "Low",
    },
    "forecast_volatility": {
        "demand_multiplier": 1.22,
        "forecast_volatility": 0.30,
        "material_cost_index_delta": -0.10,
        "supply_risk_override": "Low",
    },
    "viral_demand": {
        "demand_multiplier": 1.40,
        "demand_growth_delta": 0.16,
        "late_surge_round": 7,
        "late_surge_multiplier": 1.55,
        "material_cost_index_delta": -0.08,
        "supply_risk_override": "Low",
    },
    "premium_expansion": {
        "demand_multiplier": 1.55,
        "premium_share_shift": 0.24,
        "mid_share_shift": 0.08,
        "beginner_share_shift": -0.32,
        "quality_sensitivity_delta": 0.32,
        "premium_tech_adoption_delta": 0.34,
        "material_cost_index_delta": -0.08,
    },
}


CALIBRATED_STRATEGY_MODIFIERS: dict[str, dict[str, dict[str, Any]]] = {
    "picky_customers": {
        "premium_quality": {
            "price_multiplier": 1.14,
            "production_multiplier": 1.03,
            "forecast_multiplier": 1.00,
            "qc_delta": 3.0,
            "raw_material_multiplier": 1.02,
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 1_000.0,
        },
        "cash_conservative": {"production_multiplier": 0.68, "forecast_multiplier": 0.74},
    },
    "price_sensitive": {
        "low_cost_volume": {
            "price_multiplier": 0.94,
            "production_multiplier": 1.10,
            "forecast_multiplier": 1.02,
            "qc_delta": 0.2,
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 500.0,
            "raw_material_multiplier": 1.06,
        },
        "cash_conservative": {"production_multiplier": 0.62, "forecast_multiplier": 0.68},
    },
    "tech_shift_20": {
        "innovation_leap": {
            "price_multiplier": 1.24,
            "production_multiplier": 1.12,
            "forecast_multiplier": 1.03,
            "qc_delta": 2.2,
            "raw_material_multiplier": 1.06,
            "supplier_mix": (5.0, 35.0, 60.0),
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 1_500.0,
            "testing_intensity": 0.90,
        },
        "cash_conservative": {"production_multiplier": 0.52, "forecast_multiplier": 0.62},
    },
    "beginner_boom": {
        "low_cost_volume": {
            "price_multiplier": 0.96,
            "production_multiplier": 1.14,
            "forecast_multiplier": 1.02,
            "qc_delta": 0.1,
            "capacity_expansion_pct_of_capacity": 0.02,
            "planned_borrowing": 0.0,
            "npd_investment": 500.0,
            "raw_material_multiplier": 1.08,
        },
        "aggressive_growth": {
            "price_multiplier": 0.98,
            "production_multiplier": 1.08,
            "capacity_expansion_pct_of_capacity": 0.04,
            "planned_borrowing": 2_000.0,
        },
        "cash_conservative": {"production_multiplier": 0.62, "forecast_multiplier": 0.68},
    },
    "premium_market": {
        "premium_quality": {
            "price_multiplier": 1.18,
            "production_multiplier": 1.05,
            "qc_delta": 3.2,
            "raw_material_multiplier": 1.02,
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 1_250.0,
        },
        "cash_conservative": {"production_multiplier": 0.66, "forecast_multiplier": 0.72},
    },
    "demand_boom_20": {
        "aggressive_growth": {
            "price_multiplier": 1.02,
            "production_multiplier": 1.10,
            "forecast_multiplier": 1.06,
            "capacity_expansion_pct_of_capacity": 0.05,
            "planned_borrowing": 2_500.0,
            "npd_investment": 1_000.0,
            "raw_material_multiplier": 1.08,
        },
        "balanced_sop": {
            "production_multiplier": 1.08,
            "forecast_multiplier": 1.02,
            "capacity_expansion_pct_of_capacity": 0.03,
            "planned_borrowing": 0.0,
        },
        "cash_conservative": {"production_multiplier": 0.58, "forecast_multiplier": 0.64},
    },
    "fast_lifecycle": {
        "innovation_leap": {
            "price_multiplier": 1.22,
            "production_multiplier": 1.10,
            "forecast_multiplier": 1.03,
            "qc_delta": 2.1,
            "raw_material_multiplier": 1.06,
            "supplier_mix": (5.0, 35.0, 60.0),
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 1_500.0,
            "testing_intensity": 0.92,
        },
        "cash_conservative": {"production_multiplier": 0.52, "forecast_multiplier": 0.62},
    },
    "warranty_sensitive": {
        "premium_quality": {
            "price_multiplier": 1.16,
            "production_multiplier": 1.02,
            "qc_delta": 3.4,
            "supplier_mix": (0.0, 20.0, 80.0),
            "planned_borrowing": 0.0,
            "npd_investment": 750.0,
        },
        "cash_conservative": {"production_multiplier": 0.68, "forecast_multiplier": 0.74},
    },
    "price_war": {
        "low_cost_volume": {
            "price_multiplier": 0.95,
            "production_multiplier": 1.16,
            "forecast_multiplier": 1.03,
            "qc_delta": 0.1,
            "capacity_expansion_pct_of_capacity": 0.02,
            "planned_borrowing": 0.0,
            "npd_investment": 500.0,
            "raw_material_multiplier": 1.08,
        },
        "cash_conservative": {"production_multiplier": 0.58, "forecast_multiplier": 0.64},
    },
    "capacity_constraint": {
        "aggressive_growth": {
            "price_multiplier": 1.02,
            "production_multiplier": 1.12,
            "forecast_multiplier": 1.08,
            "capacity_expansion_pct_of_capacity": 0.05,
            "planned_borrowing": 2_500.0,
            "npd_investment": 750.0,
            "raw_material_multiplier": 1.08,
        },
        "balanced_sop": {"production_multiplier": 1.08, "planned_borrowing": 0.0},
        "cash_conservative": {"production_multiplier": 0.56, "forecast_multiplier": 0.62},
    },
    "forecast_volatility": {
        "balanced_sop": {
            "price_multiplier": 1.04,
            "production_multiplier": 1.02,
            "forecast_multiplier": 0.98,
            "capacity_expansion_pct_of_capacity": 0.01,
            "planned_borrowing": 0.0,
            "npd_investment": 1_000.0,
            "raw_material_multiplier": 1.00,
        },
        "cash_conservative": {"production_multiplier": 0.64, "forecast_multiplier": 0.68},
    },
    "viral_demand": {
        "aggressive_growth": {
            "price_multiplier": 1.03,
            "production_multiplier": 1.14,
            "forecast_multiplier": 1.08,
            "capacity_expansion_pct_of_capacity": 0.05,
            "planned_borrowing": 3_000.0,
            "npd_investment": 750.0,
            "raw_material_multiplier": 1.10,
        },
        "cash_conservative": {"production_multiplier": 0.56, "forecast_multiplier": 0.62},
    },
    "premium_expansion": {
        "premium_quality": {
            "price_multiplier": 1.18,
            "production_multiplier": 1.06,
            "qc_delta": 3.2,
            "raw_material_multiplier": 1.02,
            "capacity_expansion_pct_of_capacity": 0.02,
            "planned_borrowing": 0.0,
            "npd_investment": 1_250.0,
        },
        "cash_conservative": {"production_multiplier": 0.62, "forecast_multiplier": 0.68},
    },
}


def main() -> None:
    """Run all environments and write the workbook."""
    parser = argparse.ArgumentParser(
        description="Run 20 named offline market environments and export Excel."
    )
    parser.add_argument("--teams", type=int, default=6)
    parser.add_argument("--rounds", type=int, default=12)
    parser.add_argument(
        "--calibrated",
        action="store_true",
        help=(
            "Use scenario-specific offline calibration so different strategies "
            "have realistic chances in their favorable environments."
        ),
    )
    args = parser.parse_args()
    mode = MODE_CALIBRATED if args.calibrated else MODE_BASELINE

    if args.teams != 6:
        print("Note: this experiment is designed for the six named strategy archetypes.")

    output_dir = OUTPUT_ROOT / f"twenty_environment_{mode}_{datetime.now():%Y%m%d_%H%M%S}"
    output_dir.mkdir(parents=True, exist_ok=True)

    base_strategy_presets = deepcopy(runner.STRATEGY_PRESETS)
    runner.market_for_round = market_for_round_with_environment
    scenarios: list[dict[str, Any]] = []

    try:
        for index, base_scenario_config in enumerate(SCENARIO_DEFINITIONS, start=1):
            scenario_config = scenario_config_for_mode(base_scenario_config, mode)
            runner.STRATEGY_PRESETS = strategy_presets_for_scenario(
                scenario_config,
                mode,
                base_strategy_presets,
            )
            print(f"[{index}/20] Running {scenario_config['label']}...")
            scenarios.append(
                runner.run_strategy_scenario(
                    team_count=args.teams,
                    round_count=args.rounds,
                    strategy_rotation=STRATEGY_ROTATION,
                    market_scenario_key=scenario_config["key"],
                    market_scenario_config=scenario_config,
                    run_id=index,
                    strategy_offset=0,
                    archetype_offset=0,
                )
            )
    finally:
        runner.STRATEGY_PRESETS = base_strategy_presets

    workbook_path = output_dir / "pickleball_20_environment_simulations.xlsx"
    workbook_tables = build_workbook_tables(scenarios)
    write_excel_workbook(workbook_path, workbook_tables)
    write_csv_outputs(output_dir, workbook_tables)

    profit_counts = Counter(
        row["profit_winner_strategy"] for row in workbook_tables["all_scenario_winners"]
    )
    balanced_counts = Counter(
        row["balanced_winner_strategy"] for row in workbook_tables["all_scenario_winners"]
    )
    service_counts = Counter(
        row["service_winner_strategy"] for row in workbook_tables["all_scenario_winners"]
    )
    forecast_counts = Counter(
        row["forecast_winner_strategy"] for row in workbook_tables["all_scenario_winners"]
    )
    cash_counts = Counter(
        row["cash_winner_strategy"] for row in workbook_tables["all_scenario_winners"]
    )
    print(f"Output folder: {output_dir}")
    print(f"Excel workbook: {workbook_path}")
    print_winner_counts("Profit winner counts", profit_counts)
    print_winner_counts("Balanced winner counts", balanced_counts)
    print_winner_counts("Service winner counts", service_counts)
    print_winner_counts("Forecast winner counts", forecast_counts)
    print_winner_counts("Cash winner counts", cash_counts)

    dominance_threshold = len(SCENARIO_DEFINITIONS) * 0.60
    if profit_counts and profit_counts.most_common(1)[0][1] > dominance_threshold:
        print(
            "Calibration warning: one strategy dominates profit too much. "
            "Consider strengthening scenario-specific rewards."
        )
    if balanced_counts and balanced_counts.most_common(1)[0][1] > dominance_threshold:
        print(
            "Calibration warning: one strategy dominates balanced score too much. "
            "Consider adjusting balanced score weights or scenario-specific rewards."
        )


def scenario_config_for_mode(config: dict[str, Any], mode: str) -> dict[str, Any]:
    """Return the scenario configuration used for this offline run mode."""
    scenario = deepcopy(config)
    scenario["experiment_mode"] = mode
    scenario["applied_parameter_notes"] = (
        "The offline runner applies demand, growth, segment-share, material-cost, "
        "supply-risk, quality-sensitivity, technology, tech-adoption, beginner "
        "price-pressure, forecast-volatility, customer-pickiness, price-war, and "
        "cash-stress parameters through the existing simulator market report path."
    )
    scenario["documented_not_engine_native"] = (
        "Service sensitivity, inventory holding pressure, warranty sensitivity, "
        "and lifecycle speed are represented through supported service/fill-rate, "
        "material-cost, quality-sensitivity, technology, and demand-pressure fields "
        "rather than new engine constants."
    )
    scenario["calibrated_strategy_logic"] = "None; baseline mode uses existing strategy presets."

    if mode != MODE_CALIBRATED:
        return scenario

    overrides = CALIBRATED_SCENARIO_OVERRIDES.get(scenario["key"], {})
    for key, value in overrides.items():
        scenario[key] = value

    scenario["calibrated_strategy_logic"] = (
        "Calibrated mode keeps the real engine but adjusts offline strategy presets "
        "by scenario so strategies are not identical across worlds. For example, "
        "innovation invests and scales more in Tech Shift, premium quality leans "
        "into QC in picky/premium markets, low-cost volume scales in price-sensitive "
        "markets, and cash conservative keeps its advantage in liquidity/supply shocks."
    )
    return scenario


def strategy_presets_for_scenario(
    config: dict[str, Any],
    mode: str,
    base_strategy_presets: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    """Return strategy presets for a scenario without mutating the real app code."""
    strategy_presets = deepcopy(base_strategy_presets)
    if mode != MODE_CALIBRATED:
        return strategy_presets

    for strategy_key, updates in CALIBRATED_STRATEGY_MODIFIERS.get(config["key"], {}).items():
        if strategy_key not in strategy_presets:
            continue
        strategy_presets[strategy_key].update(updates)
    return strategy_presets


def print_winner_counts(title: str, counts: Counter) -> None:
    """Print one winner-count block in a compact, readable way."""
    print(f"{title}:")
    if not counts:
        print("- none")
        return
    for strategy, count in counts.most_common():
        print(f"- {strategy}: {count}")


def market_for_round_with_environment(
    round_number: int,
    market_scenario_key: str = "baseline",
    market_scenario_config: dict[str, Any] | None = None,
):
    """Create a deterministic market path and apply scenario-specific volatility."""
    scenario = market_scenario_config or SCENARIO_DEFINITIONS[0]
    report = BASE_MARKET_FOR_ROUND(round_number, market_scenario_key, scenario)

    volatility = float(scenario.get("forecast_volatility", 0.0))
    if volatility:
        wave = math.sin(
            (round_number + scenario["scenario_id"] * 0.37) * 1.41
        )
        report.total_demand = max(
            1,
            int(round(report.total_demand * (1.0 + volatility * wave))),
        )

    late_surge_round = scenario.get("late_surge_round")
    if late_surge_round and round_number >= int(late_surge_round):
        report.total_demand = int(
            round(report.total_demand * float(scenario.get("late_surge_multiplier", 1.0)))
        )

    customer_pickiness = float(scenario.get("customer_pickiness", 1.0))
    report.quality_sensitivity = _clamp(
        report.quality_sensitivity * customer_pickiness,
        0.0,
        1.0,
    )

    price_war_pressure = float(scenario.get("price_war_pressure", 0.0))
    price_sensitivity = float(scenario.get("price_sensitivity", 1.0))
    report.beginner_price_pressure = _clamp(
        report.beginner_price_pressure
        + 0.08 * price_war_pressure
        + 0.05 * (price_sensitivity - 1.0),
        0.0,
        1.0,
    )

    cash_stress_factor = float(scenario.get("cash_stress_factor", 1.0))
    report.material_cost_index = round(
        report.material_cost_index * (1.0 + 0.08 * max(cash_stress_factor - 1.0, 0.0)),
        4,
    )
    return report


def build_workbook_tables(scenarios: list[dict[str, Any]]) -> dict[str, Any]:
    """Build all workbook sheet payloads."""
    scenario_tables = []
    assumption_rows = []
    winner_rows = []
    teaching_rows = []
    all_final_rows = []
    all_round_rows = []

    for scenario in scenarios:
        config = scenario["market_scenario_config"]
        assumption_row = scenario_assumption_row(scenario)
        final_rows = final_team_rows_for_scenario(scenario)
        round_rows = round_detail_rows_for_scenario(scenario)
        winner_row = winner_summary_row_for_scenario(scenario, final_rows)
        teaching_row = teaching_insight_row(config, winner_row)

        scenario_tables.append(
            {
                "sheet_name": config["sheet_name"],
                "title": f"{config['sheet_name']} - {config['label']}",
                "assumptions": key_value_rows(assumption_row),
                "winners": [winner_row],
                "final_rows": final_rows,
                "round_rows": round_rows,
            }
        )
        assumption_rows.append(assumption_row)
        winner_rows.append(winner_row)
        teaching_rows.append(teaching_row)
        all_final_rows.extend(final_rows)
        all_round_rows.extend(round_rows)

    strategy_rows = strategy_summary_rows(all_final_rows, winner_rows)

    return {
        "scenario_tables": scenario_tables,
        "scenario_assumptions": assumption_rows,
        "all_scenario_winners": winner_rows,
        "strategy_summary": strategy_rows,
        "teaching_insights": teaching_rows,
        "round_level_results": all_round_rows,
        "final_team_results": all_final_rows,
    }


def scenario_assumption_row(scenario: dict[str, Any]) -> dict[str, Any]:
    """Return one assumptions row for a scenario."""
    config = scenario["market_scenario_config"]
    first_report = market_for_round_with_environment(1, config["key"], config)
    shares = first_report.normalized_shares()
    return {
        "scenario_id": config["scenario_id"],
        "scenario_name": config["label"],
        "sheet_name": config["sheet_name"],
        "experiment_mode": config.get("experiment_mode", MODE_BASELINE),
        "environment_represents": config["represents"],
        "demand_multiplier": config["demand_multiplier"],
        "demand_growth_rate": round(BASE_DEMAND_GROWTH_RATE + config["demand_growth_delta"], 4),
        "beginner_share": round(shares["beginner"], 4),
        "mid_market_share": round(shares["mid"], 4),
        "premium_share": round(shares["premium"], 4),
        "price_sensitivity": config["price_sensitivity"],
        "quality_sensitivity": round(first_report.quality_sensitivity, 4),
        "technology_shift_pressure": round(first_report.technology_shift_rate, 4),
        "supply_risk": first_report.supply_risk,
        "material_cost_multiplier": round(first_report.material_cost_index, 4),
        "forecast_volatility": config["forecast_volatility"],
        "customer_pickiness": config["customer_pickiness"],
        "price_war_pressure": config["price_war_pressure"],
        "cash_stress_factor": config["cash_stress_factor"],
        "service_sensitivity": round(config.get("service_sensitivity", config["customer_pickiness"]), 4),
        "inventory_holding_cost_pressure": round(
            config.get("inventory_holding_cost_pressure", config["cash_stress_factor"]),
            4,
        ),
        "warranty_sensitivity": round(
            config.get("warranty_sensitivity", first_report.quality_sensitivity),
            4,
        ),
        "lifecycle_speed": round(
            config.get("lifecycle_speed", first_report.technology_shift_rate),
            4,
        ),
        "market_generation": first_report.current_market_generation,
        "premium_tech_adoption": round(first_report.premium_tech_adoption, 4),
        "mid_market_tech_adoption": round(first_report.mid_market_tech_adoption, 4),
        "beginner_price_pressure": round(first_report.beginner_price_pressure, 4),
        "late_surge_round": config.get("late_surge_round", ""),
        "late_surge_multiplier": config.get("late_surge_multiplier", ""),
        "applied_parameter_notes": config.get("applied_parameter_notes", ""),
        "documented_not_engine_native": config.get("documented_not_engine_native", ""),
        "calibrated_strategy_logic": config.get("calibrated_strategy_logic", ""),
    }


def final_team_rows_for_scenario(scenario: dict[str, Any]) -> list[dict[str, Any]]:
    """Return final team rows with all requested ranks."""
    config = scenario["market_scenario_config"]
    final_results = runner.latest_round_results(scenario["team_results"])
    assignments = runner.assignment_lookup(scenario)
    innovation_scores = innovation_scores_for_scenario(scenario)

    histories_by_team = {
        result.team_name: [
            item for item in scenario["team_results"] if item.team_name == result.team_name
        ]
        for result in final_results
    }
    total_profit = {
        team_name: sum(item.profit for item in history)
        for team_name, history in histories_by_team.items()
    }
    total_revenue = {
        team_name: sum(item.revenue for item in history)
        for team_name, history in histories_by_team.items()
    }
    total_cost = {
        team_name: sum(item.total_cost for item in history)
        for team_name, history in histories_by_team.items()
    }
    service_level = {
        team_name: sum(item.sales_units for item in history)
        / max(sum(item.demand_allocated for item in history), 1.0)
        for team_name, history in histories_by_team.items()
    }
    forecast_accuracy = {
        team_name: max(
            0.0,
            1.0
            - sum(item.absolute_forecast_error_units for item in history)
            / max(sum(item.total_actual_demand_units for item in history), 1.0),
        )
        for team_name, history in histories_by_team.items()
    }
    cash_position = {
        item.team_name: item.ending_cash_balance - item.short_term_debt_balance
        for item in final_results
    }

    profit_ranks = rank_map(total_profit, descending=True)
    service_ranks = rank_map(service_level, descending=True)
    forecast_ranks = rank_map(forecast_accuracy, descending=True)
    cash_ranks = rank_map(cash_position, descending=True)
    innovation_ranks = rank_map(innovation_scores, descending=True)
    balanced_scores = {
        result.team_name: round(
            max(
                0.0,
                100.0
                * (
                    0.40
                    * runner.percentile_score(
                        total_profit[result.team_name],
                        list(total_profit.values()),
                    )
                    + 0.20
                    * runner.percentile_score(
                        service_level[result.team_name],
                        list(service_level.values()),
                    )
                    + 0.20
                    * runner.percentile_score(
                        forecast_accuracy[result.team_name],
                        list(forecast_accuracy.values()),
                    )
                    + 0.20
                    * runner.percentile_score(
                        cash_position[result.team_name],
                        list(cash_position.values()),
                    )
                    - (0.08 if result.liquidity_stress_flag else 0.0)
                ),
            ),
            1,
        )
        for result in final_results
    }
    balanced_ranks = rank_map(balanced_scores, descending=True)

    rows = []
    for result in sorted(final_results, key=lambda item: profit_ranks[item.team_name]):
        assignment = assignments[result.team_name]
        rows.append(
            {
                "experiment_mode": config.get("experiment_mode", MODE_BASELINE),
                "scenario_id": config["scenario_id"],
                "scenario_name": config["label"],
                "team_name": result.team_name,
                "strategy_archetype": assignment["strategy_label"],
                "operating_archetype": assignment["archetype"],
                "final_profit": round(total_profit[result.team_name], 2),
                "final_revenue": round(total_revenue[result.team_name], 2),
                "final_cost": round(total_cost[result.team_name], 2),
                "final_cash": round(result.ending_cash_balance, 2),
                "final_debt": round(result.short_term_debt_balance, 2),
                "cash_minus_debt": round(cash_position[result.team_name], 2),
                "ending_inventory": result.ending_inventory,
                "ending_backlog": result.backlog_units_end,
                "service_level": round(service_level[result.team_name], 4),
                "forecast_accuracy": round(forecast_accuracy[result.team_name], 4),
                "forecast_wape": round(1.0 - forecast_accuracy[result.team_name], 4),
                "balanced_score": balanced_scores[result.team_name],
                "profit_rank": profit_ranks[result.team_name],
                "balanced_rank": balanced_ranks[result.team_name],
                "service_rank": service_ranks[result.team_name],
                "forecast_rank": forecast_ranks[result.team_name],
                "cash_rank": cash_ranks[result.team_name],
                "innovation_rank": innovation_ranks[result.team_name],
                "innovation_score": round(innovation_scores[result.team_name], 2),
                "liquidity_stress": result.liquidity_stress_flag,
                "reputation_after_round": round(result.reputation_after_round, 2),
            }
        )
    return rows


def round_detail_rows_for_scenario(scenario: dict[str, Any]) -> list[dict[str, Any]]:
    """Return one row per team per round with detailed engine outputs."""
    config = scenario["market_scenario_config"]
    assignments = runner.assignment_lookup(scenario)
    rows = []
    for result in sorted(
        scenario["team_results"],
        key=lambda item: (item.round_number, item.team_name),
    ):
        assignment = assignments[result.team_name]
        rows.append(
            {
                "experiment_mode": config.get("experiment_mode", MODE_BASELINE),
                "scenario_id": config["scenario_id"],
                "scenario_name": config["label"],
                "round": result.round_number,
                "team_name": result.team_name,
                "strategy_archetype": assignment["strategy_label"],
                "operating_archetype": assignment["archetype"],
                "demand": round(result.demand_allocated, 2),
                "actual_demand": round(result.total_actual_demand_units, 2),
                "forecast": result.total_forecast_units,
                "planned_production": result.planned_production_units,
                "actual_production": result.actual_production_units,
                "units_sold": result.sales_units,
                "lost_sales": result.lost_sales_units,
                "inventory": result.ending_inventory,
                "backlog": result.backlog_units_end,
                "revenue": round(result.revenue, 2),
                "cost": round(result.total_cost, 2),
                "profit": round(result.profit, 2),
                "cash": round(result.ending_cash_balance, 2),
                "debt": round(result.short_term_debt_balance, 2),
                "service_level": round(result.fill_rate, 4),
                "forecast_error": round(result.forecast_error_units, 2),
                "forecast_wape": round(result.forecast_wape, 4),
                "defects": round(result.defect_rate, 4),
                "warranty_cost": round(result.warranty_cost, 2),
                "capacity_used": round(result.utilization_pct, 4),
                "effective_capacity": result.effective_capacity_units,
                "raw_material_inventory_end": result.ending_raw_material_inventory,
                "innovation_investment": round(result.innovation_investment, 2),
                "npd_investment": round(result.innovation_investment, 2),
                "launched_products": result.launched_project_count,
                "retired_products": result.retired_product_count,
                "launch_events": result.launch_events_text,
                "interest_expense": round(result.interest_expense, 2),
                "working_capital_requirement": round(result.working_capital_requirement, 2),
                "liquidity_stress": result.liquidity_stress_flag,
                "reputation": round(result.reputation_after_round, 2),
                "notes": result.notes,
            }
        )
    return rows


def winner_summary_row_for_scenario(
    scenario: dict[str, Any],
    final_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Return one winner-summary row for a scenario."""
    by_team = {row["team_name"]: row for row in final_rows}
    profit_winner = min(final_rows, key=lambda row: row["profit_rank"])
    balanced_winner = min(final_rows, key=lambda row: row["balanced_rank"])
    service_winner = min(final_rows, key=lambda row: row["service_rank"])
    forecast_winner = min(final_rows, key=lambda row: row["forecast_rank"])
    cash_winner = min(final_rows, key=lambda row: row["cash_rank"])
    innovation_winner = min(final_rows, key=lambda row: row["innovation_rank"])

    explanation = explain_winner(
        scenario["market_scenario_config"],
        profit_winner,
        balanced_winner,
        by_team,
    )
    return {
        "scenario_id": scenario["market_scenario_config"]["scenario_id"],
        "scenario_name": scenario["market_scenario_config"]["label"],
        "experiment_mode": scenario["market_scenario_config"].get(
            "experiment_mode",
            MODE_BASELINE,
        ),
        "profit_winner": profit_winner["team_name"],
        "profit_winner_strategy": profit_winner["strategy_archetype"],
        "profit_winner_profit": profit_winner["final_profit"],
        "balanced_winner": balanced_winner["team_name"],
        "balanced_winner_strategy": balanced_winner["strategy_archetype"],
        "balanced_score": balanced_winner["balanced_score"],
        "service_winner": service_winner["team_name"],
        "service_winner_strategy": service_winner["strategy_archetype"],
        "forecast_winner": forecast_winner["team_name"],
        "forecast_winner_strategy": forecast_winner["strategy_archetype"],
        "cash_winner": cash_winner["team_name"],
        "cash_winner_strategy": cash_winner["strategy_archetype"],
        "innovation_winner": innovation_winner["team_name"],
        "innovation_winner_strategy": innovation_winner["strategy_archetype"],
        "short_interpretation": explanation,
    }


def teaching_insight_row(
    config: dict[str, Any],
    winner_row: dict[str, Any],
) -> dict[str, Any]:
    """Return a teaching insight row for a scenario."""
    best_strategy = str(winner_row["balanced_winner_strategy"])
    return {
        "scenario_id": config["scenario_id"],
        "scenario_name": config["label"],
        "experiment_mode": config.get("experiment_mode", MODE_BASELINE),
        "what_this_environment_represents": config["represents"],
        "best_strategy": best_strategy,
        "why_it_won": winner_row["short_interpretation"],
        "what_students_should_learn": config["learning"],
        "discussion_question_for_class": config["discussion_question"],
    }


def strategy_summary_rows(
    final_rows: list[dict[str, Any]],
    winner_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Aggregate performance across all 20 environments."""
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in final_rows:
        grouped[str(row["strategy_archetype"])].append(row)

    profit_wins = Counter(row["profit_winner_strategy"] for row in winner_rows)
    balanced_wins = Counter(row["balanced_winner_strategy"] for row in winner_rows)
    service_wins = Counter(row["service_winner_strategy"] for row in winner_rows)
    forecast_wins = Counter(row["forecast_winner_strategy"] for row in winner_rows)
    cash_wins = Counter(row["cash_winner_strategy"] for row in winner_rows)
    innovation_wins = Counter(row["innovation_winner_strategy"] for row in winner_rows)

    rows = []
    for strategy, strategy_rows_for_group in grouped.items():
        count = max(len(strategy_rows_for_group), 1)
        rows.append(
            {
                "experiment_mode": strategy_rows_for_group[0].get(
                    "experiment_mode",
                    MODE_BASELINE,
                ),
                "strategy_archetype": strategy,
                "profit_wins": profit_wins[strategy],
                "balanced_wins": balanced_wins[strategy],
                "service_wins": service_wins[strategy],
                "forecast_wins": forecast_wins[strategy],
                "cash_wins": cash_wins[strategy],
                "innovation_wins": innovation_wins[strategy],
                "average_final_profit": round(
                    sum(row["final_profit"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_final_revenue": round(
                    sum(row["final_revenue"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_final_cash": round(
                    sum(row["final_cash"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_final_debt": round(
                    sum(row["final_debt"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_service_level": round(
                    sum(row["service_level"] for row in strategy_rows_for_group) / count,
                    4,
                ),
                "average_forecast_accuracy": round(
                    sum(row["forecast_accuracy"] for row in strategy_rows_for_group) / count,
                    4,
                ),
                "average_balanced_score": round(
                    sum(row["balanced_score"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_profit_rank": round(
                    sum(row["profit_rank"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "average_balanced_rank": round(
                    sum(row["balanced_rank"] for row in strategy_rows_for_group) / count,
                    2,
                ),
                "main_strength": STRATEGY_STRENGTHS.get(strategy, ""),
                "main_weakness": STRATEGY_WEAKNESSES.get(strategy, ""),
            }
        )
    return sorted(rows, key=lambda row: (row["balanced_wins"], row["profit_wins"]), reverse=True)


def innovation_scores_for_scenario(scenario: dict[str, Any]) -> dict[str, float]:
    """Score innovation position using cumulative investment, launches, and tech level."""
    final_results = runner.latest_round_results(scenario["team_results"])
    scores: dict[str, float] = {}
    for final_result in final_results:
        team_history = [
            item for item in scenario["team_results"] if item.team_name == final_result.team_name
        ]
        total_innovation = sum(item.innovation_investment for item in team_history)
        total_launches = sum(item.launched_project_count for item in team_history)
        total_ready = sum(item.launch_ready_project_count for item in team_history)
        scores[final_result.team_name] = (
            final_result.average_portfolio_tech_generation * 50.0
            + total_launches * 12.0
            + total_ready * 4.0
            + total_innovation / 2_000.0
        )
    return scores


def explain_winner(
    config: dict[str, Any],
    profit_winner: dict[str, Any],
    balanced_winner: dict[str, Any],
    by_team: dict[str, dict[str, Any]],
) -> str:
    """Create a short scenario-specific interpretation."""
    strategy = str(profit_winner["strategy_archetype"])
    if strategy == "Cash conservative":
        return (
            "Cash conservative protected liquidity and avoided debt while other teams absorbed "
            "cost, inventory, expansion, or innovation pressure."
        )
    if strategy == "Balanced S&OP":
        return (
            "Balanced S&OP aligned forecasts, production, and service well enough to convert "
            "demand into profit without the highest risk profile."
        )
    if strategy == "Premium quality":
        return (
            "Premium quality benefited because this environment rewarded reliability, service, "
            "and reputation enough to offset higher operating cost."
        )
    if strategy == "Innovation leap":
        return (
            "Innovation leap benefited from technology pressure and future-product readiness, "
            "but its cash burden should still be inspected."
        )
    if strategy == "Aggressive growth":
        return (
            "Aggressive growth captured demand upside, but the class should inspect whether "
            "expansion and debt made the result fragile."
        )
    if strategy == "Low-cost volume":
        return (
            "Low-cost volume found enough price-sensitive demand to offset its thinner margins "
            "and operating risk."
        )
    return (
        f"{strategy} won profit in a {config['label']} environment; compare its cash, "
        f"service, forecast accuracy, and inventory position against {balanced_winner['strategy_archetype']}."
    )


def rank_map(values_by_team: dict[str, float], descending: bool) -> dict[str, int]:
    """Return competition ranks from a team-to-value mapping."""
    sorted_items = sorted(
        values_by_team.items(),
        key=lambda item: item[1],
        reverse=descending,
    )
    return {team_name: index + 1 for index, (team_name, _) in enumerate(sorted_items)}


def key_value_rows(row: dict[str, Any]) -> list[dict[str, Any]]:
    """Convert an assumptions row into a key-value table."""
    return [{"parameter": key, "value": value} for key, value in row.items()]


def write_csv_outputs(output_dir: Path, tables: dict[str, Any]) -> None:
    """Write flat CSV companion files for researchers who prefer spreadsheets/dataframes."""
    csv_payloads = {
        "scenario_assumptions.csv": tables["scenario_assumptions"],
        "all_scenario_winners.csv": tables["all_scenario_winners"],
        "strategy_summary.csv": tables["strategy_summary"],
        "teaching_insights.csv": tables["teaching_insights"],
        "round_level_results.csv": tables["round_level_results"],
        "final_team_results.csv": tables["final_team_results"],
    }
    for file_name, rows in csv_payloads.items():
        write_csv_file(output_dir / file_name, rows)


def write_csv_file(path: Path, rows: list[dict[str, Any]]) -> None:
    """Write a CSV with stable union-of-keys headers."""
    if not rows:
        path.write_text("", encoding="utf-8")
        return

    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)

    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: safe_csv_value(row.get(header)) for header in headers})


def safe_csv_value(value: Any) -> Any:
    """Convert values into CSV-friendly scalars."""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if value is None:
        return ""
    return value


def write_excel_workbook(path: Path, tables: dict[str, Any]) -> None:
    """Write all simulation outputs into a formatted Excel workbook."""
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    add_simple_table_sheet(
        workbook,
        "Scenario_Assumptions",
        "Scenario Assumptions",
        tables["scenario_assumptions"],
        "ScenarioAssumptions",
    )
    add_simple_table_sheet(
        workbook,
        "All_Scenario_Winners",
        "All Scenario Winners",
        tables["all_scenario_winners"],
        "AllScenarioWinners",
    )
    add_simple_table_sheet(
        workbook,
        "Strategy_Summary",
        "Strategy Summary",
        tables["strategy_summary"],
        "StrategySummary",
    )
    add_simple_table_sheet(
        workbook,
        "Teaching_Insights",
        "Teaching Insights",
        tables["teaching_insights"],
        "TeachingInsights",
    )

    for scenario_table in tables["scenario_tables"]:
        add_scenario_sheet(workbook, scenario_table)

    workbook.save(path)


def add_simple_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    title: str,
    rows: list[dict[str, Any]],
    table_name: str,
) -> None:
    """Create one summary sheet with one formatted table."""
    sheet = workbook.create_sheet(sheet_name)
    sheet.freeze_panes = "A2"
    write_title(sheet, title, 1, max(len(rows[0]) if rows else 1, 4))
    end_row = write_table(sheet, 3, rows, table_name)
    apply_sheet_formatting(sheet, max_row=end_row)


def add_scenario_sheet(workbook: Workbook, scenario_table: dict[str, Any]) -> None:
    """Create one scenario-specific sheet with four sections."""
    sheet = workbook.create_sheet(scenario_table["sheet_name"])
    sheet.freeze_panes = "A2"
    write_title(sheet, scenario_table["title"], 1, 12)

    row = 3
    row = write_section(sheet, row, "D. Scenario Winner Summary")
    row = write_table(
        sheet,
        row,
        scenario_table["winners"],
        f"T{scenario_table['sheet_name'][:2]}Winners",
    ) + 2

    row = write_section(sheet, row, "A. Scenario Assumptions")
    row = write_table(
        sheet,
        row,
        scenario_table["assumptions"],
        f"T{scenario_table['sheet_name'][:2]}Assumptions",
    ) + 2

    row = write_section(sheet, row, "B. Final Team Results")
    row = write_table(
        sheet,
        row,
        scenario_table["final_rows"],
        f"T{scenario_table['sheet_name'][:2]}Final",
    ) + 2

    row = write_section(sheet, row, "C. Round-by-Round Details")
    end_row = write_table(
        sheet,
        row,
        scenario_table["round_rows"],
        f"T{scenario_table['sheet_name'][:2]}Rounds",
    )

    apply_sheet_formatting(sheet, max_row=end_row)


def write_title(sheet, title: str, row: int, span_columns: int) -> None:
    """Write a merged title row."""
    sheet.cell(row=row, column=1, value=title)
    sheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=max(span_columns, 1),
    )
    cell = sheet.cell(row=row, column=1)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_section(sheet, row: int, title: str) -> int:
    """Write a section label and return the next row."""
    sheet.cell(row=row, column=1, value=title)
    cell = sheet.cell(row=row, column=1)
    cell.font = Font(bold=True, size=12, color="1F4E78")
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    return row + 1


def write_table(sheet, start_row: int, rows: list[dict[str, Any]], table_name: str) -> int:
    """Write a list of dictionaries as a formatted Excel table."""
    if not rows:
        sheet.cell(row=start_row, column=1, value="No rows")
        return start_row

    headers = list(rows[0])
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row in enumerate(rows, start=1):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(
                row=start_row + row_offset,
                column=column_index,
                value=safe_excel_value(row.get(header)),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            apply_number_format(cell, header)

    end_row = start_row + len(rows)
    end_column = len(headers)
    table_ref = (
        f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    )
    table = Table(displayName=clean_table_name(table_name), ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    return end_row


def apply_sheet_formatting(sheet, max_row: int) -> None:
    """Apply widths and row styling to a worksheet."""
    for row in sheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value is not None:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                cell.alignment = alignment

    for column_index in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row_index in range(1, max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        width = min(max(max_length + 2, 10), 42)
        sheet.column_dimensions[column_letter].width = width

    for row_number in range(1, max_row + 1):
        sheet.row_dimensions[row_number].height = 20


def apply_number_format(cell, header: str) -> None:
    """Format financial, percentage, and rank columns."""
    normalized = header.lower()
    if any(
        token in normalized
        for token in (
            "profit",
            "revenue",
            "cost",
            "cash",
            "debt",
            "warranty",
            "interest",
            "capital",
        )
    ) and "rank" not in normalized and "score" not in normalized:
        cell.number_format = '$#,##0.00;[Red]-$#,##0.00'
    elif any(
        token in normalized
        for token in ("service_level", "forecast_accuracy", "forecast_wape", "defects", "capacity_used")
    ):
        cell.number_format = "0.0%"
    elif "score" in normalized:
        cell.number_format = "0.0"
    elif "rank" in normalized or "round" in normalized or "units" in normalized:
        cell.number_format = "0"
    elif isinstance(cell.value, float):
        cell.number_format = "0.00"


def clean_table_name(name: str) -> str:
    """Return an Excel-safe table name."""
    cleaned = "".join(character for character in name if character.isalnum() or character == "_")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:240]


def safe_excel_value(value: Any) -> Any:
    """Convert Python values into Excel-friendly values."""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if value is None:
        return ""
    return value


def _clamp(value: float, minimum: float, maximum: float) -> float:
    """Clamp a value into a closed range."""
    return max(minimum, min(value, maximum))


if __name__ == "__main__":
    main()
